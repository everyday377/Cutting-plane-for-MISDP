# -*- coding: utf-8 -*-
"""
Created on Sun Mar 29 18:46:15 2020

@author: wangyuzhu
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Jan 20 19:18:45 2020

@author: wangyuzhu
"""
from gurobipy import *
import scipy.io
import numpy as np
import time
from openpyxl import Workbook
from scipy.sparse import csr_matrix, csc_matrix, coo_matrix, lil_matrix

wb=Workbook()


def mycallback(model,where):
    
    if where==GRB.Callback.MIPSOL:
        n=list(x.keys())[-1][0]+1
        x1=np.zeros((n,n))
        for i in range(n):
            for j in range(i,n):
                x1[i,j]=model.cbGetSolution(x[i,j])
                x1[j,i]=model.cbGetSolution(x[i,j])
        w,v=np.linalg.eig(x1)
        if np.amin(w)<-tol:#tol
            vmin=np.zeros((n,1))
            vmin[:,0]=v[:,np.argmin(w)]
            pmin=np.matmul(vmin,np.transpose(vmin))
            model.cbLazy(quicksum(2*pmin[i,j]*x[i,j] for (i,j) in x)-quicksum(pmin[i,i]*x[i,i] for i in range(n))>=0)




ne=[30,10,10,10]# size of matrix A in R^(n*m), m=10
me=[10,30,30,50]# m the constraint number
zne=[10,10,30,10]# the number of integer variables


for ii in range(1): #range(2) kappa
    
    n=ne[ii]
    m=me[ii]
    zn=zne[ii]
    C=scipy.io.loadmat('MISDP_xn'+str(n)+'_m'+str(m)+'_zn'+str(zn)+'.mat')['MISDP'][0,0]['C']
    Ax={}
    Az={}
    for i in range(m):
        Ax[i]={}
        Ax[i]=scipy.io.loadmat('MISDP_xn'+str(n)+'_m'+str(m)+'_zn'+str(zn)+'.mat')['MISDP'][0,0]['A'][0,i]['Ax']
        Az[i]={}
        Az[i]=scipy.io.loadmat('MISDP_xn'+str(n)+'_m'+str(m)+'_zn'+str(zn)+'.mat')['MISDP'][0,0]['A'][0,i]['az']
    b=scipy.io.loadmat('MISDP_xn'+str(n)+'_m'+str(m)+'_zn'+str(zn)+'.mat')['MISDP'][0,0]['b']
    zl=scipy.io.loadmat('MISDP_xn'+str(n)+'_m'+str(m)+'_zn'+str(zn)+'.mat')['MISDP'][0,0]['zl']
    zu=scipy.io.loadmat('MISDP_xn'+str(n)+'_m'+str(m)+'_zn'+str(zn)+'.mat')['MISDP'][0,0]['zu']
    tol=1e-5*(1+sum(abs(b[0])));
    sense='min'
    ws=wb.create_sheet('result_BAC_xn'+str(n)+'_m'+str(m)+'_zn'+str(zn))
            
    for ll in range(2): #range(3) outer approximation of the semidefinite cone
                
        if ll==0:
            method='DD'
        elif ll==1:
            method='SDB'
                
        vec=scipy.io.loadmat('Apx_'+method+'_n2_'+str(n)+'.mat')['vec']
        vecr,vecc=vec.shape
            
            
            
                # create MIP model
        tstart=time.time()
        model=Model()
        x={}
        z={}
        u={}
        for i in range(n):
            for j in range(i,n):#i<=j
                x[i,j]=model.addVar(vtype='C',lb=-GRB.INFINITY)#define variables x
        for i in range(zn):
            z[i]=model.addVar(lb=zl[i,0],ub=zu[i,0],vtype='I')
                    
        model.update()
                #set objective function
        model.setObjective(quicksum(2*C[i,j]*x[i,j] for (i,j) in x)-quicksum(C[i,i]*x[i,i] for i in range(n)),GRB.MINIMIZE)

                
                #constraints
                #Axx+Azz=b
        for k in range(m):
            model.addConstr(quicksum(2*Ax[k][i,j]*x[i,j] for (i,j) in x)
                                            -quicksum(Ax[k][i,i]*x[i,i] for i in range(n))
                                            +quicksum(Az[k][i,0]*z[i] for i in range(zn))==b[0][k])
                #<X,vec>>=0
        for k in range(vecc):
            model.addConstr(quicksum(2*vec[n*i+j,k]*x[i,j] for (i,j) in x)-quicksum(vec[n*i+i,k]*x[i,i] for i in range(n))>=0)
            
        model.Params.DualReductions=0
        model.Params.LazyConstraints=1
        model.Params.LogToConsole=0
        model.Params.LogFile='MISDP_xn'+str(n)+'_m'+str(m)+'_zn'+str(zn)+'_3.26.txt'
                
                
                
        tmodel=time.time()-tstart
                
                
        model.optimize(mycallback)
                
        obj=model.ObjVal
        tcal=model.Runtime
        nodes=model.getAttr('NodeCount')
        
                #np.savez('result_RIC_'+str(n)+'_'+str(kappa)+'_'+'_1.24.npz',obj,tmodel,tcal)
                
                
                
                
                
        ws['A1']='Method'
        ws['A'+str(ll+2)]=method
        ws['B1']='Optimal value'
        ws['B'+str(ll+2)]=obj
        ws['C1']='Run time'
        ws['C'+str(ll+2)]=tcal
        ws['D1']='Model time'
        ws['D'+str(ll+2)]=tmodel
        ws['E1']='Nodes'
        ws['E'+str(ll+2)]=nodes
                
                
wb.save('result_MISDP_total_3.26.xlsx')
